"""Parse de planilha CRM BSOD (xlsx / CSV / HTML Excel)."""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from datetime import datetime
from typing import Any

from bs4 import BeautifulSoup
from openpyxl import load_workbook

from lib.util import as_str, normalize_vlan

CRM_FIELD_KEYS = (
    "protocolo",
    "uf",
    "svlan",
    "cvlan",
    "contrato_netsms",
    "cadastro_responsavel",
    "cliente",
    "tipo_logradouro",
    "logradouro",
    "numero",
    "complemento",
    "bairro",
    "cep",
    "cidade",
    "produto",
    "designacao",
    "status",
    "contato_cliente_nome_1",
    "contato_cliente_telefone_1",
    "contato_cliente_nome_2",
    "contato_cliente_telefone_2",
    "contato_cliente_email_1",
    "contato_cliente_email_2",
    "contrato_conectado",
    "construcao_data_execucao",
    "cancelamento_data",
    "cancelamento_motivo",
)

_HEADER_ALIASES = {
    "protocolo": "protocolo",
    "proto": "protocolo",
    "uf": "uf",
    "svlan": "svlan",
    "s_vlan": "svlan",
    "cvlan": "cvlan",
    "c_vlan": "cvlan",
    "contrato_netsms": "contrato_netsms",
    "contrato_net_sms": "contrato_netsms",
    "cadastro_responsavel": "cadastro_responsavel",
    "responsavel_cadastro": "cadastro_responsavel",
    "cliente": "cliente",
    "nome_cliente": "cliente",
    "tipo_logradouro": "tipo_logradouro",
    "tipo_do_logradouro": "tipo_logradouro",
    "logradouro": "logradouro",
    "endereco": "logradouro",
    "numero": "numero",
    "num": "numero",
    "complemento": "complemento",
    "bairro": "bairro",
    "cep": "cep",
    "cidade": "cidade",
    "produto": "produto",
    "designacao": "designacao",
    "status": "status",
    "situacao": "status",
    "contato_cliente_nome_1": "contato_cliente_nome_1",
    "contato_nome_1": "contato_cliente_nome_1",
    "contato_cliente_telefone_1": "contato_cliente_telefone_1",
    "contato_telefone_1": "contato_cliente_telefone_1",
    "contato_cliente_nome_2": "contato_cliente_nome_2",
    "contato_nome_2": "contato_cliente_nome_2",
    "contato_cliente_telefone_2": "contato_cliente_telefone_2",
    "contato_telefone_2": "contato_cliente_telefone_2",
    "contato_cliente_email_1": "contato_cliente_email_1",
    "contato_email_1": "contato_cliente_email_1",
    "contato_cliente_email_2": "contato_cliente_email_2",
    "contato_email_2": "contato_cliente_email_2",
    "contrato_conectado": "contrato_conectado",
    "construcao_data_execucao": "construcao_data_execucao",
    "data_execucao": "construcao_data_execucao",
    "cancelamento_data": "cancelamento_data",
    "data_cancelamento": "cancelamento_data",
    "cancelamento_motivo": "cancelamento_motivo",
    "motivo_cancelamento": "cancelamento_motivo",
}


def _normalize_header(raw: Any) -> str:
    """Normaliza cabeçalho de coluna para chave canônica."""
    text = as_str(raw, max_len=120)
    if not text:
        return ""
    nfkd = unicodedata.normalize("NFKD", text)
    ascii_text = "".join(ch for ch in nfkd if not unicodedata.combining(ch))
    key = re.sub(r"[^a-z0-9]+", "_", ascii_text.lower()).strip("_")
    return _HEADER_ALIASES.get(key, key)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return as_str(value, max_len=255)


def _map_header_row(headers: list[Any]) -> list[str]:
    return [_normalize_header(h) for h in headers]


def _row_from_values(headers: list[str], values: list[Any]) -> dict[str, str] | None:
    """Monta dict CRM; None se protocolo ausente."""
    mapped: dict[str, str] = {key: "" for key in CRM_FIELD_KEYS}
    for idx, header in enumerate(headers):
        if header not in mapped:
            continue
        mapped[header] = _cell_text(values[idx] if idx < len(values) else "")
    mapped["protocolo"] = as_str(mapped.get("protocolo"), max_len=64)
    mapped["contrato_netsms"] = as_str(mapped.get("contrato_netsms"), max_len=64)
    mapped["cvlan"] = normalize_vlan(mapped.get("cvlan"))
    mapped["svlan"] = normalize_vlan(mapped.get("svlan"))
    mapped["uf"] = as_str(mapped.get("uf"), max_len=8).upper()
    if not mapped["protocolo"]:
        return None
    return mapped


def _require_headers(headers: list[str], kind: str) -> None:
    """Exige coluna protocolo na planilha."""
    if not headers or "protocolo" not in headers:
        raise RuntimeError(f"Planilha {kind} sem coluna protocolo (headers={headers[:20]})")


def _parse_xlsx(payload: bytes) -> list[dict[str, str]]:
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        headers = _map_header_row(list(next(rows_iter, []) or []))
        _require_headers(headers, "xlsx")
        out: list[dict[str, str]] = []
        for raw in rows_iter:
            row = _row_from_values(headers, list(raw or []))
            if row:
                out.append(row)
        return out
    finally:
        workbook.close()


def _parse_csv(payload: bytes) -> list[dict[str, str]]:
    text = payload.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
    except csv.Error:
        dialect = csv.excel
        dialect.delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), dialect)
    headers = _map_header_row(next(reader, []))
    _require_headers(headers, "CSV")
    out: list[dict[str, str]] = []
    for raw in reader:
        row = _row_from_values(headers, list(raw or []))
        if row:
            out.append(row)
    return out


def _largest_table(soup: BeautifulSoup) -> Any:
    tables = soup.find_all("table")
    if not tables:
        return None
    return max(tables, key=lambda t: len(t.find_all("tr")))


def _parse_html_table(payload: bytes) -> list[dict[str, str]]:
    soup = BeautifulSoup(payload, "html.parser")
    table = _largest_table(soup)
    if table is None:
        raise RuntimeError("Planilha HTML sem tabela")
    trs = table.find_all("tr")
    if len(trs) < 2:
        raise RuntimeError("Planilha HTML sem linhas de dados")
    headers = _map_header_row([cell.get_text(" ", strip=True) for cell in trs[0].find_all(["th", "td"])])
    if "protocolo" not in headers:
        headers = _map_header_row(
            [cell.get_text(" ", strip=True) for cell in trs[1].find_all(["th", "td"])]
        )
        data_trs = trs[2:]
    else:
        data_trs = trs[1:]
    _require_headers(headers, "HTML")
    out: list[dict[str, str]] = []
    for tr in data_trs:
        values = [cell.get_text(" ", strip=True) for cell in tr.find_all(["td", "th"])]
        row = _row_from_values(headers, values)
        if row:
            out.append(row)
    return out


def parse_planilha(payload: bytes, content_type: str = "") -> list[dict[str, str]]:
    """Detecta formato da planilha e retorna linhas CRM."""
    if not payload:
        return []
    ctype = (content_type or "").lower()
    if payload[:2] == b"PK":
        return _parse_xlsx(payload)
    if b"<html" in payload[:2000].lower() or b"<table" in payload[:4000].lower() or "html" in ctype:
        return _parse_html_table(payload)
    if "excel" in ctype or "spreadsheet" in ctype:
        try:
            return _parse_xlsx(payload)
        except Exception:
            return _parse_html_table(payload)
    if b"," in payload[:512] or b";" in payload[:512] or "csv" in ctype or "text/" in ctype:
        return _parse_csv(payload)
    try:
        return _parse_html_table(payload)
    except Exception:
        return _parse_csv(payload)


def is_crm_cancelled(row: dict[str, str]) -> bool:
    """Indica se a linha CRM tem STATUS cancelado (não entra no catálogo)."""
    status = as_str(row.get("status"), max_len=64)
    if not status:
        return False
    nfkd = unicodedata.normalize("NFKD", status)
    ascii_text = "".join(ch for ch in nfkd if not unicodedata.combining(ch))
    compact = re.sub(r"[^a-z0-9]+", "", ascii_text.lower())
    return compact == "cancelado" or compact.endswith("cancelado")


def exclude_cancelled_crm_rows(rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], int]:
    """Remove linhas STATUS=CANCELADO; retorna (ativas, quantidade_excluida)."""
    active: list[dict[str, str]] = []
    skipped = 0
    for row in rows:
        if is_crm_cancelled(row):
            skipped += 1
            continue
        active.append(row)
    return active, skipped


def dedupe_by_protocolo(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Mantém uma linha por protocolo (última ganha)."""
    by_proto: dict[str, dict[str, str]] = {}
    for row in rows:
        proto = row.get("protocolo") or ""
        if not proto:
            continue
        by_proto[proto] = row
    return list(by_proto.values())


# Compatibilidade com imports antigos do módulo nocclaro.
def dedupe_by_cvlan(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Deprecated: preferir dedupe_by_protocolo."""
    return dedupe_by_protocolo(rows)
