"""Parse de planilha local BSOD (cliente, designação, VLAN, contrato)."""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from lib.util import as_str, normalize_contrato

SHEET_FIELD_KEYS = (
    "cliente",
    "designacao",
    "razao_social",
    "endereco",
    "contrato_netsms",
    "vlan",
)

_HEADER_ALIASES = {
    "cliente": "cliente",
    "designacao": "designacao",
    "designacao_": "designacao",
    "razao_social": "razao_social",
    "razao": "razao_social",
    "endereco": "endereco",
    "contrato": "contrato_netsms",
    "contrato_netsms": "contrato_netsms",
    "contrato_net_sms": "contrato_netsms",
    "vlan": "vlan",
    "cvlan": "vlan",
    "c_vlan": "vlan",
}


def _normalize_header(raw: Any) -> str:
    """Normaliza cabeçalho de coluna para chave canônica da planilha local."""
    text = as_str(raw, max_len=120)
    if not text:
        return ""
    nfkd = unicodedata.normalize("NFKD", text)
    ascii_text = "".join(ch for ch in nfkd if not unicodedata.combining(ch))
    key = re.sub(r"[^a-z0-9]+", "_", ascii_text.lower()).strip("_")
    return _HEADER_ALIASES.get(key, key)


def _cell_text(value: Any) -> str:
    """Converte célula da planilha para texto; inteiros sem sufixo .0."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return as_str(value, max_len=255)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return as_str(value, max_len=255)


def _row_from_values(headers: list[str], values: list[Any]) -> dict[str, str] | None:
    """Monta linha da planilha local; None se contrato ausente."""
    mapped: dict[str, str] = {key: "" for key in SHEET_FIELD_KEYS}
    for idx, header in enumerate(headers):
        if header not in mapped:
            continue
        mapped[header] = _cell_text(values[idx] if idx < len(values) else "")
    mapped["contrato_netsms"] = as_str(mapped.get("contrato_netsms"), max_len=64)
    if not mapped["contrato_netsms"]:
        return None
    return mapped


def contrato_keys(raw: Any) -> list[str]:
    """Extrai um ou mais contratos normalizados; trata par ``123 / 456``."""
    text = as_str(raw, max_len=64)
    if not text:
        return []
    if re.search(r"\d{6,}\s*/\s*\d{6,}", text):
        keys: list[str] = []
        for part in re.split(r"\s*/\s*", text):
            key = normalize_contrato(part)
            if key:
                keys.append(key)
        return keys
    key = normalize_contrato(text)
    return [key] if key else []


def parse_local_sheet(path: Path) -> list[dict[str, str]]:
    """Lê xlsx local e retorna linhas deduplicadas por contrato normalizado."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        headers = [_normalize_header(h) for h in list(next(rows_iter, []) or [])]
        if "contrato_netsms" not in headers:
            raise RuntimeError(f"Planilha sem coluna CONTRATO (headers={headers})")
        parsed: list[dict[str, str]] = []
        for raw in rows_iter:
            row = _row_from_values(headers, list(raw or []))
            if row:
                parsed.append(row)
    finally:
        workbook.close()
    return dedupe_by_contrato(parsed)


def dedupe_by_contrato(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Mantém uma linha por contrato normalizado (última ganha)."""
    by_contrato: dict[str, dict[str, str]] = {}
    for row in rows:
        for key in contrato_keys(row.get("contrato_netsms")):
            by_contrato[key] = row
    seen: set[int] = set()
    unique: list[dict[str, str]] = []
    for row in by_contrato.values():
        marker = id(row)
        if marker in seen:
            continue
        seen.add(marker)
        unique.append(row)
    return unique
